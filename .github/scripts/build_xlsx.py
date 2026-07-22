#!/usr/bin/env python3
"""Genera data/citas.xlsx a partir de data/citas.json.

Columnas (en el orden pedido): Fecha y hora, Vamos a, Horario, Me apetece,
Ubicación, Sitio. Se ejecuta dentro de la GitHub Action tras append.js.
"""

import json
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

DATA_DIR = "data"
JSON_PATH = os.path.join(DATA_DIR, "citas.json")
XLSX_PATH = os.path.join(DATA_DIR, "citas.xlsx")

# (encabezado visible, clave en el JSON)
COLUMNS = [
    ("Fecha y hora", "fecha"),
    ("Vamos a", "plan"),
    ("Horario", "franja"),
    ("Me apetece", "antojo"),
    ("Ubicación", "ciudad"),
    ("Sitio", "sitio"),
]


def load_rows():
    try:
        with open(JSON_PATH, encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, list) else []
    except (FileNotFoundError, json.JSONDecodeError):
        return []


def fmt_fecha(iso):
    if not iso:
        return ""
    try:
        # ISO con Z -> datetime, formato es-ES
        dt = datetime.fromisoformat(str(iso).replace("Z", "+00:00"))
        return dt.strftime("%d/%m/%Y %H:%M")
    except ValueError:
        return str(iso)


def main():
    os.makedirs(DATA_DIR, exist_ok=True)
    rows = load_rows()

    wb = Workbook()
    ws = wb.active
    ws.title = "Citas"

    header_fill = PatternFill("solid", fgColor="EF5B47")
    header_font = Font(bold=True, color="FFFFFF")

    # Encabezados
    for col, (title, _) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=col, value=title)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(vertical="center")

    # Datos
    for i, rec in enumerate(rows, start=2):
        for col, (_, key) in enumerate(COLUMNS, start=1):
            val = rec.get(key, "") if isinstance(rec, dict) else ""
            if key == "fecha":
                val = fmt_fecha(val)
            ws.cell(row=i, column=col, value=val)

    # Ancho de columnas
    widths = [20, 14, 14, 16, 18, 30]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(COLUMNS)).column_letter}{max(1, len(rows) + 1)}"

    wb.save(XLSX_PATH)
    print(f"Excel generado con {len(rows)} filas -> {XLSX_PATH}")


if __name__ == "__main__":
    main()
